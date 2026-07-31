from datetime import datetime

from openpyxl import load_workbook

from concertowl.export_excel import export_excel
from concertowl.models import PriceSnapshot
from concertowl.repo_history import RepoHistory
from concertowl.storage import RepoStorage
from concertowl.view_trends import build_payload, write_dashboard
from concertowl.watchlist import WATCHLIST_HEADER


def add_run(
    root,
    run_id,
    day,
    price,
    *,
    show_datetime="2026-08-07T19:30",
    source="piaoniu",
):
    history = RepoHistory(root, run_id=run_id, now=day)
    history.record(
        [
            PriceSnapshot(
                observed_at=day.isoformat(timespec="minutes"),
                event_id="event-1",
                artist="薛之谦",
                city="重庆",
                tour="天外来物",
                show_datetime=show_datetime,
                face_price=517,
                observed_price=price,
                currency="CNY",
                source=source,
                status="在售",
                note="tier=517",
            )
        ]
    )
    history.finalize({"status": "success", "sources": {"piaoniu": {"snapshots": 1}}})


def write_watchlist(root, *, show_datetime="2026-08-07T19:30"):
    storage = RepoStorage(root)
    storage.overwrite(
        "Watchlist",
        [
            WATCHLIST_HEADER,
            [
                "event-1",
                "薛之谦",
                "天外来物",
                "重庆",
                "大陆",
                "",
                show_datetime,
                "",
                "",
                "",
                "",
                "",
                "auto",
                "true",
            ],
        ],
    )


def test_dashboard_and_excel_smoke(tmp_path):
    add_run(tmp_path, "one", datetime(2026, 7, 30, 8), 600)
    add_run(tmp_path, "two", datetime(2026, 7, 31, 8), 580)
    write_watchlist(tmp_path)

    payload = build_payload(tmp_path)
    assert payload["event_count"] == 1
    assert payload["events"][0]["delta"] == -20
    index = write_dashboard(payload, tmp_path / "site")
    html = index.read_text(encoding="utf-8")
    assert "最近运行" in html
    assert 'window.addEventListener("pageshow"' in html
    assert "eventSel.value=events.some" in html

    workbook_path = export_excel(tmp_path, tmp_path / "report.xlsx")
    workbook = load_workbook(workbook_path, read_only=True)
    assert {"概览", "Watchlist", "运行记录", "价_薛之谦"} <= set(
        workbook.sheetnames
    )


def test_dashboard_rejects_observations_from_another_session(tmp_path):
    add_run(
        tmp_path,
        "matching",
        datetime(2026, 7, 30, 8),
        600,
        show_datetime="2026-08-07T19:30",
        source="moretickets",
    )
    add_run(
        tmp_path,
        "wrong-session",
        datetime(2026, 7, 31, 8),
        999,
        show_datetime="2026-08-08T19:30",
        source="piaoniu",
    )
    write_watchlist(tmp_path, show_datetime="2026-08-07T19:30")

    payload = build_payload(tmp_path)

    assert payload["event_count"] == 1
    event = payload["events"][0]
    assert event["show"] == "2026-08-07T19:30"
    assert event["last"] == 600
    assert event["sources"] == ["moretickets"]
