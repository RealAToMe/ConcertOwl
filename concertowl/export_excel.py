"""Export repository history to an Excel workbook on demand."""
from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from .repo_history import data_root_from_env, iter_observations, iter_run_manifests
from .storage import RepoStorage

OBSERVATION_COLUMNS = [
    "observed_at",
    "event_id",
    "artist",
    "city",
    "tour",
    "show_datetime",
    "face_price",
    "observed_price",
    "premium_ratio",
    "days_to_show",
    "days_since_onsale",
    "currency",
    "source",
    "status",
    "record_kind",
    "collect_run_id",
    "note",
]


def _sheet_title(value: str, used: set[str]) -> str:
    title = "价_" + "".join(ch for ch in value if ch not in "[]:*?/\\")
    title = title[:31] or "价格"
    base, counter = title, 2
    while title in used:
        suffix = f"_{counter}"
        title = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(title)
    return title


def _filter_rows(
    data_root: Path,
    artist: str = "",
    start: str = "",
    end: str = "",
) -> list[dict]:
    rows = []
    for row in iter_observations(data_root):
        observed_at = str(row.get("observed_at") or "")
        if artist and row.get("artist") != artist:
            continue
        if start and observed_at[:10] < start:
            continue
        if end and observed_at[:10] > end:
            continue
        rows.append(row)
    return rows


def export_excel(
    data_root: Path | str,
    output: Path | str,
    *,
    artist: str = "",
    start: str = "",
    end: str = "",
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    root = Path(data_root)
    destination = Path(output)
    destination.parent.mkdir(parents=True, exist_ok=True)
    observations = _filter_rows(root, artist, start, end)
    runs = sorted(
        iter_run_manifests(root),
        key=lambda run: str(run.get("completed_at") or ""),
        reverse=True,
    )

    workbook = Workbook()
    overview = workbook.active
    overview.title = "概览"
    by_artist = Counter(str(row.get("artist") or "") for row in observations)
    by_source = Counter(str(row.get("source") or "") for row in observations)
    overview.append(["项目", "值"])
    overview.append(["导出时间", datetime.now().isoformat(timespec="seconds")])
    overview.append(["筛选歌手", artist or "全部"])
    overview.append(["日期范围", f"{start or '不限'} ~ {end or '不限'}"])
    overview.append(["观测记录", len(observations)])
    overview.append(["歌手数", len(by_artist)])
    overview.append(["运行批次", len(runs)])
    overview.append([])
    overview.append(["歌手", "记录数"])
    for name, count in by_artist.most_common():
        overview.append([name, count])
    overview.append([])
    overview.append(["来源", "记录数"])
    for source, count in by_source.most_common():
        overview.append([source, count])

    storage = RepoStorage(root)
    watch_rows = storage.read_rows("Watchlist")
    watch = workbook.create_sheet("Watchlist")
    for row in watch_rows:
        watch.append(row)

    run_sheet = workbook.create_sheet("运行记录")
    run_columns = [
        "run_id",
        "started_at",
        "completed_at",
        "status",
        "events_total",
        "records_seen",
        "records_written",
        "changes",
        "heartbeats",
        "unchanged",
    ]
    run_sheet.append(run_columns)
    for run in runs:
        run_sheet.append([run.get(column, "") for column in run_columns])

    grouped = defaultdict(list)
    for row in observations:
        grouped[str(row.get("artist") or "未知")].append(row)
    used = {"概览", "Watchlist", "运行记录"}
    for name in sorted(grouped):
        sheet = workbook.create_sheet(_sheet_title(name, used))
        sheet.append(OBSERVATION_COLUMNS)
        for row in sorted(
            grouped[name],
            key=lambda item: (
                str(item.get("event_id") or ""),
                str(item.get("observed_at") or ""),
                str(item.get("face_price") or ""),
            ),
        ):
            sheet.append([row.get(column, "") for column in OBSERVATION_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column in range(1, min(sheet.max_column, 18) + 1):
            width = 12
            for cell in list(sheet.columns)[column - 1][:100]:
                width = max(width, min(42, len(str(cell.value or "")) + 2))
            sheet.column_dimensions[get_column_letter(column)].width = width

    workbook.save(destination)
    return destination


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir")
    parser.add_argument("--output")
    parser.add_argument("--artist", default="")
    parser.add_argument("--start", default="", help="YYYY-MM-DD")
    parser.add_argument("--end", default="", help="YYYY-MM-DD")
    args = parser.parse_args(argv)
    root = Path(args.data_dir) if args.data_dir else data_root_from_env()
    output = (
        Path(args.output)
        if args.output
        else Path("exports")
        / f"ConcertOwl_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    path = export_excel(
        root, output, artist=args.artist, start=args.start, end=args.end
    )
    print(f"[excel] {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
